from aiogram import Router, types, F
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.filters import StateFilter
from sqlalchemy import select
import openpyxl
import io
from datetime import date, time, datetime, timedelta
from bot.db.database import async_session
from bot.db.models import Group, Schedule, Poll
from bot.logger import logger

router = Router()

cancel_kb = types.InlineKeyboardMarkup(
    inline_keyboard=[[types.InlineKeyboardButton(text="❌ Отмена", callback_data="cancel_action")]]
)

class ScheduleUpload(StatesGroup):
    waiting_for_file = State()

def extract_type(discipline: str) -> str:
    """Извлекает тип занятия: lecture, practice, lab, other."""
    d = discipline.strip().lower()
    if d.startswith("лек"):
        return "lecture"
    if d.startswith("пр"):
        return "practice"
    if d.startswith("лаб"):
        return "lab"
    return "other"

# --- Обработчик выбора группы (для админа) ---
@router.callback_query(F.data == "admin_upload_schedule")
async def admin_upload_schedule_choose(callback: types.CallbackQuery, state: FSMContext):
    async with async_session() as session:
        groups = (await session.execute(select(Group))).scalars().all()
        if not groups:
            await callback.message.answer("Нет групп.")
            await callback.answer()
            return
        keyboard = types.InlineKeyboardMarkup(inline_keyboard=[
            [types.InlineKeyboardButton(text=g.name, callback_data=f"schedgroup_{g.id}")] for g in groups
        ] + [[types.InlineKeyboardButton(text="❌ Отмена", callback_data="cancel_action")]])
    await callback.message.answer("Выберите группу для загрузки расписания:", reply_markup=keyboard)
    await callback.answer()

# --- Обработчик выбора группы (для reply-кнопки, если пользователь староста или админ) ---
@router.callback_query(F.data.startswith("schedgroup_"))
async def choose_group_for_schedule(callback: types.CallbackQuery, state: FSMContext):
    group_id = int(callback.data.split("_")[1])
    async with async_session() as session:
        group = await session.get(Group, group_id)
        if not group:
            await callback.message.answer("Группа не найдена.")
            await callback.answer()
            return
        await state.update_data(group_id=group.id, group_name=group.name)
    await callback.message.answer(
        f"Отправьте Excel-файл с расписанием для группы {group.name}.\n"
        "Ожидаемые столбцы: День недели, Дата, Время с, Время по, Дисциплина, Преподаватель, Аудитория, Группа",
        reply_markup=cancel_kb
    )
    await state.set_state(ScheduleUpload.waiting_for_file)
    await callback.answer()

# --- Обработка получения файла ---
@router.message(StateFilter(ScheduleUpload.waiting_for_file), F.document)
async def process_schedule_file(message: types.Message, state: FSMContext):
    document = message.document
    if not document.file_name.endswith(".xlsx"):
        await message.answer("Файл должен быть в формате .xlsx")
        return

    data = await state.get_data()
    group_id = data["group_id"]

    buf = io.BytesIO()
    await message.bot.download(document, destination=buf)
    buf.seek(0)

    try:
        wb = openpyxl.load_workbook(buf, read_only=True)
        ws = wb.active
    except Exception as e:
        logger.error(f"Excel read error: {e}")
        await message.answer("Ошибка чтения файла. Проверьте формат.")
        await state.clear()
        return

    # Определяем индексы столбцов по названиям
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        date_col = headers.index("Дата")
        start_col = headers.index("Время с")
        end_col = headers.index("Время по")
        disc_col = headers.index("Дисциплина")
        teacher_col = headers.index("Преподаватель")
        audience_col = headers.index("Аудитория")
    except ValueError:
        await message.answer("Не найдены обязательные столбцы (Дата, Время с, Время по, Дисциплина, Преподаватель, Аудитория).")
        await state.clear()
        return

    # Читаем строки
    raw_lessons = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not all([row[date_col], row[start_col], row[end_col], row[disc_col]]):
            continue
        try:
            if isinstance(row[date_col], date):
                lesson_date = row[date_col]
            else:
                lesson_date = datetime.strptime(str(row[date_col]), "%d.%m.%Y").date()
            start_time = row[start_col]
            if not isinstance(start_time, time):
                start_time = datetime.strptime(str(start_time), "%H:%M").time()
            end_time = row[end_col]
            if not isinstance(end_time, time):
                end_time = datetime.strptime(str(end_time), "%H:%M").time()
        except Exception:
            continue
        disc = str(row[disc_col]).strip()
        teacher = str(row[teacher_col]).strip() if row[teacher_col] else ""
        audience = str(row[audience_col]).strip() if row[audience_col] else ""
        raw_lessons.append((lesson_date, start_time, end_time, disc, teacher, audience))

    if not raw_lessons:
        await message.answer("Файл не содержит корректных записей о занятиях.")
        await state.clear()
        return

    # Объединение сдвоенных пар
    raw_lessons.sort(key=lambda x: (x[0], x[1]))
    merged = []
    for lesson in raw_lessons:
        if not merged:
            merged.append(list(lesson))
            continue
        last = merged[-1]
        # Условия объединения: та же дата, одинаковые дисциплина и тип, перерыв <=30 мин
        if (lesson[0] == last[0] and
            extract_type(lesson[3]) == extract_type(last[3]) and
            lesson[3] == last[3] and
            (datetime.combine(lesson[0], lesson[1]) - datetime.combine(last[0], last[2])).total_seconds() <= 30*60):
            last[2] = max(last[2], lesson[2])
            if lesson[4] not in last[4]:
                last[4] += ", " + lesson[4]
            if lesson[5] not in last[5]:
                last[5] += ", " + lesson[5]
        else:
            merged.append(list(lesson))

    # Сохранение в БД
    async with async_session() as session:
        new_count = 0
        updated_count = 0
        for lesson in merged:
            lesson_date, start_time, end_time, disc, teacher, audience = lesson
            existing = (await session.execute(
                select(Schedule).where(
                    Schedule.group_id == group_id,
                    Schedule.date == lesson_date,
                    Schedule.time_start == start_time,
                    Schedule.discipline == disc,
                    Schedule.type == extract_type(disc)
                )
            )).scalars().first()

            if existing:
                # Если опросов ещё нет, обновляем
                if not (await session.execute(select(Poll).where(Poll.schedule_id == existing.id))).scalars().first():
                    existing.time_end = end_time
                    existing.teacher = teacher
                    existing.audience = audience
                    updated_count += 1
            else:
                new_schedule = Schedule(
                    group_id=group_id,
                    date=lesson_date,
                    time_start=start_time,
                    time_end=end_time,
                    discipline=disc,
                    type=extract_type(disc),
                    teacher=teacher,
                    audience=audience
                )
                session.add(new_schedule)
                new_count += 1

        await session.commit()

    await message.answer(
        f"Расписание загружено.\nНовых занятий: {new_count}\nОбновлено: {updated_count}"
    )
    await state.clear()