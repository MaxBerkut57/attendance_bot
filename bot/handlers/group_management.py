from aiogram import Router, types, F
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.filters import StateFilter
from sqlalchemy import select
import openpyxl
import io
import secrets
from bot.db.database import async_session
from bot.db.models import User, Group, GroupMembership, PendingInvite
from bot.keyboards.main_menu import get_main_menu
from bot.logger import logger

router = Router()

cancel_kb = types.InlineKeyboardMarkup(
    inline_keyboard=[[types.InlineKeyboardButton(text="❌ Отмена", callback_data="cancel_action")]]
)

class GroupManagement(StatesGroup):
    waiting_for_file = State()

async def get_starosta_group(session, user_id: int) -> Group | None:
    stmt = select(Group).where(Group.starosta_id == user_id)
    result = await session.execute(stmt)
    return result.scalars().first()

# ==================== МЕНЮ СТАРОСТЫ ====================
@router.callback_query(F.data == "menu_starosta_group")
async def starosta_menu(callback: types.CallbackQuery):
    keyboard = types.InlineKeyboardMarkup(inline_keyboard=[
        [types.InlineKeyboardButton(text="📋 Загрузить список студентов", callback_data="starosta_upload_list")],
        [types.InlineKeyboardButton(text="🔙 Назад", callback_data="menu_back")]
    ])
    await callback.message.edit_text("Управление группой:", reply_markup=keyboard)
    await callback.answer()

# ==================== ЗАГРУЗКА (староста) ====================
@router.callback_query(F.data == "starosta_upload_list")
async def starosta_upload_list(callback: types.CallbackQuery, state: FSMContext):
    async with async_session() as session:
        group = await get_starosta_group(session, callback.from_user.id)
        if not group:
            await callback.message.answer("Вы не являетесь старостой ни одной группы.")
            await callback.answer()
            return
        await state.update_data(group_id=group.id, group_name=group.name)
    await callback.message.answer(
        f"Отправьте Excel-файл со списком студентов группы {group.name}.\n"
        "Формат: колонки «ФИО» и «username» (без @).",
        reply_markup=cancel_kb
    )
    await state.set_state(GroupManagement.waiting_for_file)
    await callback.answer()

# ==================== ЗАГРУЗКА (админ для любой группы) ====================
@router.callback_query(F.data == "admin_upload_list")
async def admin_upload_list_choose_group(callback: types.CallbackQuery):
    async with async_session() as session:
        groups = (await session.execute(select(Group))).scalars().all()
        if not groups:
            await callback.message.answer("Нет доступных групп.")
            await callback.answer()
            return
        keyboard = types.InlineKeyboardMarkup(inline_keyboard=[
            [types.InlineKeyboardButton(text=g.name, callback_data=f"uploadgroup_{g.id}")] for g in groups
        ] + [[types.InlineKeyboardButton(text="❌ Отмена", callback_data="cancel_action")]])
    await callback.message.answer("Выберите группу для загрузки списка:", reply_markup=keyboard)
    await callback.answer()

@router.callback_query(F.data.startswith("uploadgroup_"))
async def admin_upload_list_for_group(callback: types.CallbackQuery, state: FSMContext):
    group_id = int(callback.data.split("_")[1])
    async with async_session() as session:
        group = await session.get(Group, group_id)
        if not group:
            await callback.message.answer("Группа не найдена.")
            await callback.answer()
            return
        await state.update_data(group_id=group.id, group_name=group.name)
    await callback.message.answer(
        f"Отправьте Excel-файл со списком студентов группы {group.name}.\n"
        "Формат: колонки «ФИО» и «username» (без @).",
        reply_markup=cancel_kb
    )
    await state.set_state(GroupManagement.waiting_for_file)
    await callback.answer()

# ==================== ОБРАБОТКА ФАЙЛА ====================
@router.message(StateFilter(GroupManagement.waiting_for_file), F.document)
async def process_student_list_file(message: types.Message, state: FSMContext):
    document = message.document
    if not document.file_name.endswith(".xlsx"):
        await message.answer("Пожалуйста, отправьте файл в формате Excel (.xlsx).")
        return

    data = await state.get_data()
    group_id = data["group_id"]

    buf = io.BytesIO()
    await message.bot.download(document, destination=buf)
    buf.seek(0)

    try:
        wb = openpyxl.load_workbook(buf, read_only=True)
        ws = wb.active
    except Exception as e:
        logger.error(f"Excel read error: {e}")
        await message.answer("Ошибка чтения файла. Проверьте формат.")
        await state.clear()
        return

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    name_col = username_col = None
    for i, h in enumerate(headers):
        if h and "фио" in str(h).lower():
            name_col = i
        if h and "username" in str(h).lower():
            username_col = i

    if name_col is None or username_col is None:
        await message.answer("Не найдены колонки «ФИО» и «username».")
        await state.clear()
        return

    added = 0
    updated = 0
    errors = 0
    pending_links = []

    async with async_session() as session:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) <= max(name_col, username_col):
                continue
            full_name = str(row[name_col]).strip() if row[name_col] else ""
            username = str(row[username_col]).strip().lstrip("@") if row[username_col] else ""

            if not full_name:
                errors += 1
                continue

            if not username:
                token = secrets.token_urlsafe(32)
                artificial_username = f"invite_{token}"
                user = User(
                    user_id=None,
                    username=artificial_username,
                    full_name=full_name,
                    is_admin=False
                )
                session.add(user)
                await session.flush()
                session.add(GroupMembership(user_id=None, group_id=group_id))
                session.add(PendingInvite(token=token, user_id=user.user_id))
                pending_links.append((full_name, token))
                logger.info(f"Creating invite for {full_name}")
                continue

            stmt = select(User).where(User.username == username)
            result = await session.execute(stmt)
            user = result.scalars().first()
            if not user:
                user = User(
                    user_id=None,
                    username=username,
                    full_name=full_name,
                    is_admin=False
                )
                session.add(user)
                await session.flush()
                added += 1
            else:
                if user.full_name != full_name:
                    user.full_name = full_name
                    updated += 1

            existing_membership = await session.execute(
                select(GroupMembership).where(
                    GroupMembership.group_id == group_id,
                    GroupMembership.user.has(User.username == username)
                )
            )
            if not existing_membership.scalars().first():
                session.add(GroupMembership(
                    user_id=user.user_id,
                    group_id=group_id
                ))

        await session.commit()

    result_text = (
        f"Загрузка завершена.\n"
        f"Добавлено новых пользователей: {added}\n"
        f"Обновлено ФИО: {updated}\n"
        f"Строк с ошибками: {errors}"
    )
    if pending_links:
        bot_username = (await message.bot.get_me()).username
        links_text = "\n".join(
            f"{name}: https://t.me/{bot_username}?start=invite_{token}"
            for name, token in pending_links
        )
        result_text += f"\n\nСсылки для студентов без username:\n{links_text}"

    await message.answer(result_text)
    await state.clear()