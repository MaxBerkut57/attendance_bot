import io
import pandas as pd
from datetime import datetime, date
from sqlalchemy import select
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from bot.db.database import async_session
from bot.db.models import (Poll, Attendance, User, GroupMembership,
                           Group, Schedule, PollMessage)
from bot.logger import logger
from aiogram.types import BufferedInputFile
from bot.services.bot_instance import get_bot

# Цвета
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)

async def generate_attendance_excel(poll_id: int) -> bytes:
    """Генерирует Excel с результатами конкретного опроса (для автоматической отправки после закрытия)."""
    async with async_session() as session:
        poll = await session.get(Poll, poll_id)
        if not poll:
            raise ValueError("Опрос не найден")
        schedule = await session.get(Schedule, poll.schedule_id)
        group = await session.get(Group, schedule.group_id)

        members = (await session.execute(
            select(User).join(GroupMembership).where(GroupMembership.group_id == group.id)
        )).scalars().all()

        wb = Workbook()
        ws = wb.active
        ws.title = f"Опрос {poll.id}"
        ws.append(["ФИО", "Статус"])

        # Стили для заголовка
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT

        for user in members:
            att = await session.get(Attendance, (poll_id, user.user_id))
            if att:
                status = "присутствовал" if att.status == "present" else "отсутствовал"
                fill = GREEN_FILL if status == "присутствовал" else RED_FILL
            else:
                # Если опрос активен – "не отметился" жёлтым, иначе красным
                if poll.status == 'active':
                    status = "не отметился"
                    fill = YELLOW_FILL
                else:
                    status = "не отметился"
                    fill = RED_FILL
            row = [user.full_name, status]
            ws.append(row)
            ws.cell(row=ws._current_row, column=2).fill = fill

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

async def generate_group_report_for_date(group_id: int, sched_id: int) -> bytes:
    """Отчёт по конкретному занятию (для ручного запроса)."""
    async with async_session() as session:
        schedule = await session.get(Schedule, sched_id)
        if not schedule:
            raise ValueError("Расписание не найдено")
        poll = (await session.execute(
            select(Poll).where(Poll.schedule_id == sched_id)
        )).scalars().first()
        if not poll:
            raise ValueError("Опрос не найден")

        members = (await session.execute(
            select(User).join(GroupMembership).where(GroupMembership.group_id == group_id)
        )).scalars().all()

        wb = Workbook()
        ws = wb.active
        ws.title = f"{schedule.discipline}"
        ws.append(["ФИО", "Статус"])
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT

        for user in members:
            att = await session.get(Attendance, (poll.id, user.user_id))
            if att:
                status = "присутствовал" if att.status == "present" else "отсутствовал"
                fill = GREEN_FILL if status == "присутствовал" else RED_FILL
            else:
                if poll.status == 'active':
                    status = "не отметился"
                    fill = YELLOW_FILL
                else:
                    status = "не отметился"
                    fill = RED_FILL
            row = [user.full_name, status]
            ws.append(row)
            ws.cell(row=ws._current_row, column=2).fill = fill

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

async def generate_curator_attendance_report_for_date(group_id: int, target_date: date) -> bytes:
    """Сводный отчёт за конкретную дату для куратора."""
    async with async_session() as session:
        schedules = (await session.execute(
            select(Schedule).join(Poll).where(
                Schedule.group_id == group_id,
                Schedule.date == target_date,
                Poll.status.in_(['active', 'finished'])
            )
        )).scalars().all()

        if not schedules:
            raise ValueError("Нет занятий с опросами на эту дату")

        members = (await session.execute(
            select(User).join(GroupMembership).where(GroupMembership.group_id == group_id)
        )).scalars().all()

        wb = Workbook()
        ws = wb.active
        ws.title = f"Посещаемость {target_date}"
        headers = ["ФИО"] + [s.discipline for s in schedules]
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT

        # Для каждого студента собираем статусы
        for user in members:
            row = [user.full_name]
            for sched in schedules:
                poll = (await session.execute(
                    select(Poll).where(Poll.schedule_id == sched.id)
                )).scalars().first()
                if not poll:
                    row.append("")  # нет опроса — пусто
                    continue
                att = await session.get(Attendance, (poll.id, user.user_id))
                if att:
                    status = "+" if att.status == "present" else "-"
                else:
                    status = "?" if poll.status == 'active' else "-"
                row.append(status)
            ws.append(row)
            # Раскрашиваем ячейки статусов
            for col_idx, value in enumerate(row[1:], start=2):
                cell = ws.cell(row=ws._current_row, column=col_idx)
                if value == "+":
                    cell.fill = GREEN_FILL
                elif value == "-":
                    cell.fill = RED_FILL
                elif value == "?":
                    cell.fill = YELLOW_FILL

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

async def send_report_to_starosta(poll: Poll):
    """Отправляет отчёт старосте группы."""
    bot = get_bot()
    async with async_session() as session:
        schedule = await session.get(Schedule, poll.schedule_id)
        group = await session.get(Group, schedule.group_id)
        if not group or not group.starosta_id:
            return
        try:
            file_bytes = await generate_attendance_excel(poll.id)
            discipline = schedule.discipline.replace(" ", "_")
            date_str = schedule.date.strftime("%d.%m.%Y")
            caption = f"Отчёт: {schedule.discipline} ({date_str} {schedule.time_start}-{schedule.time_end})"
            file = BufferedInputFile(file_bytes, filename=f"{discipline}_{date_str}.xlsx")
            await bot.send_document(
                chat_id=group.starosta_id,
                document=file,
                caption=caption
            )
            poll.report_sent = True
            await session.commit()
            logger.info(f"Report sent to starosta {group.starosta_id} for poll {poll.id}")
        except Exception as e:
            logger.error(f"Failed to send report: {e}")